import time
from lxml import etree
from selenium.webdriver import Chrome
from selenium.webdriver import ChromeOptions
from selenium.webdriver.chrome.options import Options #不显示页面
import openpyxl

# 实现无可视化界面操作
chrome_options = Options()
chrome_options.add_argument('--headless')
chrome_options.add_argument('--disable-gpu')

# 实现规避检查
options = ChromeOptions()
options.add_experimental_option('excludeSwitches', ['enable-automation'])

#如何实现让 selenium 规避被检测到的风险
web = Chrome(chrome_options=chrome_options, options=options)#把参数设置到浏览器

name = input("请输入你想要搜索的商品")
# name = "apple"
count = 2

data = openpyxl.load_workbook(r'apple.xlsx')
data1 = data.active
# 目标网站：https://www.aliexpress.com/

for page in range(1,2):

    url = f'https://es.aliexpress.com/wholesale?trafficChannel=main&d=y&CatId=0&SearchText={name}&ltype=wholesale&SortType=default&page={page}'
    # url = "https://es.aliexpress.com/wholesale?catId=0&initiative_id=SB_20211016063715&SearchText=apple"
    # url = 'https://es.aliexpress.com/item/4000967199322.html?spm=a2g0o.productlist.0.0.11941e9cpeE9oE&algo_pvid=d0cbb703-bfa7-4580-bb65-fb61'

    web.get(url)
    text = web.page_source  #得到页面element的html代码
    tree = etree.HTML(text)
    data_urls = tree.xpath('//div/a[@class="awV9E"]/@href')
    for data_url in data_urls:
        url1 = "https://es.aliexpress.com" + data_url
        web.get(url1)
        text = web.page_source  #得到页面element的html代码
        tree = etree.HTML(text)
        data_title = tree.xpath('//div/h1[@class="product-title-text"]/text()')
        data_title = ''.join(data_title)

        # 将数据存入表格
        data1.cell(row=count, column=1, value=count-1)
        data1.cell(row=count, column=2, value=data_title)
        data1.cell(row=count, column=3, value=url1)
        count += 1

        # print(data_title)
        time.sleep(1)

    time.sleep(2)
web.quit()

data.save(r'apple.xlsx')

print("运行结束！")